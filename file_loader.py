import os
import logging
from typing import Dict, Any, List, Union
from pathlib import Path
from abc import ABC, abstractmethod
import json
import xml.etree.ElementTree as ET
import zipfile
import tarfile
import csv
import io

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FileHandler(ABC):
    @abstractmethod
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        pass

class TextFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                content = file.read()
            return {"type": "text", "content": content}
        except IOError as e:
            logger.error(f"Error reading text file {file_path}: {e}")
            return {"type": "text", "error": str(e)}

class CSVFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.DictReader(file)
                data = list(reader)
            return {"type": "csv", "headers": reader.fieldnames, "data": data}
        except IOError as e:
            logger.error(f"Error reading CSV file {file_path}: {e}")
            return {"type": "csv", "error": str(e)}

class PDFFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            import PyPDF2
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return {"type": "pdf", "content": text, "num_pages": len(pdf_reader.pages)}
        except ImportError:
            logger.error("PyPDF2 module not found. Skipping PDF file.")
            return {"type": "pdf", "error": "PyPDF2 module not found"}
        except Exception as e:
            logger.error(f"Error reading PDF file {file_path}: {e}")
            return {"type": "pdf", "error": str(e)}

class ImageFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            from PIL import Image
            import io
            import base64
            with Image.open(file_path) as img:
                # Convert image to base64 string
                buffered = io.BytesIO()
                img.save(buffered, format=img.format)
                img_str = base64.b64encode(buffered.getvalue()).decode()
                return {
                    "type": "image",
                    "format": img.format,
                    "mode": img.mode,
                    "size": img.size,
                    "base64": img_str
                }
        except ImportError:
            logger.error("Pillow module not found. Skipping image file.")
            return {"type": "image", "error": "Pillow module not found"}
        except Exception as e:
            logger.error(f"Error reading image file {file_path}: {e}")
            return {"type": "image", "error": str(e)}

class JSONFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
            return {"type": "json", "content": data}
        except json.JSONDecodeError as e:
            logger.error(f"Error parsing JSON file {file_path}: {e}")
            return {"type": "json", "error": str(e)}

class XMLFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            return {"type": "xml", "content": self._element_to_dict(root)}
        except ET.ParseError as e:
            logger.error(f"Error parsing XML file {file_path}: {e}")
            return {"type": "xml", "error": str(e)}

    def _element_to_dict(self, element):
        result = {}
        for child in element:
            child_data = self._element_to_dict(child)
            if child.tag in result:
                if type(result[child.tag]) is list:
                    result[child.tag].append(child_data)
                else:
                    result[child.tag] = [result[child.tag], child_data]
            else:
                result[child.tag] = child_data
        if element.text and element.text.strip():
            result['#text'] = element.text.strip()
        return result

class ZIPFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                content = {name: zip_ref.read(name).decode('utf-8', errors='ignore') for name in file_list if not name.endswith('/')}
            return {"type": "zip", "file_list": file_list, "content": content}
        except zipfile.BadZipFile as e:
            logger.error(f"Error reading ZIP file {file_path}: {e}")
            return {"type": "zip", "error": str(e)}

class TARFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            with tarfile.open(file_path, 'r:*') as tar_ref:
                file_list = tar_ref.getnames()
                content = {}
                for name in file_list:
                    member = tar_ref.getmember(name)
                    if member.isfile():
                        f = tar_ref.extractfile(member)
                        content[name] = f.read().decode('utf-8', errors='ignore')
            return {"type": "tar", "file_list": file_list, "content": content}
        except tarfile.TarError as e:
            logger.error(f"Error reading TAR file {file_path}: {e}")
            return {"type": "tar", "error": str(e)}

class HTMLFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            from bs4 import BeautifulSoup
            with open(file_path, 'r', encoding="utf-8") as file:
                soup = BeautifulSoup(file, 'html.parser')
            return {
                "type": "html",
                "title": soup.title.string if soup.title else None,
                "text": soup.get_text(),
                "links": [{"href": link.get("href"), "text": link.text} for link in soup.find_all("a")]
            }
        except ImportError:
            logger.error("BeautifulSoup module not found. Skipping HTML file.")
            return {"type": "html", "error": "BeautifulSoup module not found"}
        except Exception as e:
            logger.error(f"Error reading HTML file {file_path}: {e}")
            return {"type": "html", "error": str(e)}

class ExcelFileHandler(FileHandler):
    def extract_data(self, file_path: Path) -> Dict[str, Any]:
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            data = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data[sheet_name] = [
                    [cell.value for cell in row]
                    for row in sheet.iter_rows()
                ]
            return {"type": "excel", "sheets": data}
        except ImportError:
            logger.error("openpyxl module not found. Skipping Excel file.")
            return {"type": "excel", "error": "openpyxl module not found"}
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            return {"type": "excel", "error": str(e)}

class FileProcessor:
    def __init__(self):
        self.handlers: Dict[str, FileHandler] = {
            ".txt": TextFileHandler(),
            ".log": TextFileHandler(),
            ".csv": CSVFileHandler(),
            ".pdf": PDFFileHandler(),
            ".jpg": ImageFileHandler(),
            ".jpeg": ImageFileHandler(),
            ".png": ImageFileHandler(),
            ".gif": ImageFileHandler(),
            ".json": JSONFileHandler(),
            ".xml": XMLFileHandler(),
            ".zip": ZIPFileHandler(),
            ".tar": TARFileHandler(),
            ".gz": TARFileHandler(),
            ".html": HTMLFileHandler(),
            ".htm": HTMLFileHandler(),
            ".xlsx": ExcelFileHandler(),
            ".xls": ExcelFileHandler()
        }

    def register_handler(self, extension: str, handler: FileHandler) -> None:
        self.handlers[extension.lower()] = handler

    def process_file(self, file_path: Path) -> Dict[str, Any]:
        file_extension = file_path.suffix.lower()
        handler = self.handlers.get(file_extension)
        if handler:
            return handler.extract_data(file_path)
        else:
            logger.warning(f"No handler for file extension: {file_path}")
            return {"type": "unknown", "error": f"No handler for file extension: {file_extension}"}

    def process_directory(self, directory: Path, recursive: bool = True) -> Dict[str, Dict[str, Any]]:
        if not directory.is_dir():
            raise ValueError(f"{directory} is not a valid directory.")

        results = {}
        for entry in os.scandir(directory):
            entry_path = Path(entry)
            if entry.is_file():
                results[str(entry_path)] = self.process_file(entry_path)
            elif recursive and entry.is_dir():
                results.update(self.process_directory(entry_path))
        return results

def main() -> None:
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Process files in a directory and prepare data for LLM or other programs.")
    parser.add_argument("directory", type=str, help="Directory path to process")
    parser.add_argument("--recursive", action="store_true", help="Process subdirectories recursively")
    parser.add_argument("--log", choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'], 
                        default='INFO', help="Set the logging level")
    parser.add_argument("--output", type=str, help="Output file path for extracted data (JSON format)")

    args = parser.parse_args()

    # Set logging level based on argument
    logging.getLogger().setLevel(args.log)

    processor = FileProcessor()
    try:
        results = processor.process_directory(Path(args.directory), args.recursive)
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            logger.info(f"Results written to {args.output}")
        else:
            print(json.dumps(results, indent=2, ensure_ascii=False))
    except Exception as e:
        logger.critical(f"An error occurred: {e}")

if __name__ == "__main__":
    main()


# example_usage.py

# from pathlib import Path
# from data_extractor import FileProcessor
# import json

# def process_and_analyze(directory_path: str, recursive: bool = True):
#     # Create an instance of FileProcessor
#     processor = FileProcessor()

#     # Process the directory
#     results = processor.process_directory(Path(directory_path), recursive)

#     # Analyze the results
#     file_types = {}
#     total_files = 0
#     errors = []

#     for file_path, data in results.items():
#         total_files += 1
#         file_type = data.get('type', 'unknown')
#         file_types[file_type] = file_types.get(file_type, 0) + 1

#         if 'error' in data:
#             errors.append((file_path, data['error']))

#     # Print summary
#     print(f"Processed {total_files} files:")
#     for file_type, count in file_types.items():
#         print(f"  - {file_type}: {count}")

#     if errors:
#         print("\nErrors encountered:")
#         for file_path, error in errors:
#             print(f"  - {file_path}: {error}")

#     # Example of accessing specific data
#     for file_path, data in results.items():
#         if data['type'] == 'text':
#             print(f"\nContent preview of {file_path}:")
#             print(data['content'][:100] + "...")  # Print first 100 characters
#         elif data['type'] == 'csv':
#             print(f"\nFirst row of CSV {file_path}:")
#             print(data['data'][0] if data['data'] else "Empty CSV")

#     # Save full results to a file
#     with open('full_results.json', 'w') as f:
#         json.dump(results, f, indent=2)
#     print("\nFull results saved to full_results.json")

# if __name__ == "__main__":
#     process_and_analyze("/path/to/your/directory", recursive=True)